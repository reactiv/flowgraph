#!/usr/bin/env python3
"""Excel formula recalculation script using LibreOffice.

Recalculates all formulas in an Excel file and checks for errors.
Requires LibreOffice to be installed.

Usage:
    python recalc.py <excel_file> [timeout_seconds]
"""

import json
import os
import platform
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

# Excel error values to detect
EXCEL_ERRORS = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]


LIBRARY_NAME = "TransformerRecalc"


def get_libreoffice_macro_dir():
    """Get the LibreOffice macro directory for our custom library."""
    if platform.system() == "Darwin":
        base = Path.home() / "Library/Application Support/LibreOffice/4/user/basic"
    else:
        base = Path.home() / ".config/libreoffice/4/user/basic"
    return base / LIBRARY_NAME


def create_recalc_macro(macro_dir: Path):
    """Create a LibreOffice Basic macro for recalculation.

    Uses a dedicated library (TransformerRecalc) to avoid overwriting
    user's existing macros in the Standard library.
    """
    macro_dir.mkdir(parents=True, exist_ok=True)

    macro_code = '''Sub RecalcAndSave
    Dim oDoc As Object
    oDoc = ThisComponent
    oDoc.calculateAll()
    oDoc.store()
    oDoc.close(True)
End Sub
'''

    (macro_dir / "RecalcMacro.xba").write_text(macro_code)

    # Create script.xlb for our custom library
    script_xlb = f'''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">
<library:library xmlns:library="http://openoffice.org/2000/library" library:name="{LIBRARY_NAME}" library:readonly="false" library:passwordprotected="false">
 <library:element library:name="RecalcMacro"/>
</library:library>
'''
    (macro_dir / "script.xlb").write_text(script_xlb)


def run_libreoffice_recalc(file_path: Path, timeout: int = 60):
    """Run LibreOffice to recalculate formulas."""
    # Find LibreOffice executable
    if platform.system() == "Darwin":
        soffice = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        # On macOS, gtimeout requires coreutils (brew install coreutils)
        # Fall back to running without timeout if not available
        if shutil.which("gtimeout"):
            timeout_cmd = "gtimeout"
        else:
            timeout_cmd = None  # Will use Python's subprocess timeout instead
    else:
        soffice = "soffice"
        timeout_cmd = "timeout"

    if not shutil.which(soffice) and platform.system() != "Darwin":
        soffice = "libreoffice"

    # Create macro
    macro_dir = get_libreoffice_macro_dir()
    create_recalc_macro(macro_dir)

    # Run LibreOffice in headless mode with macro
    abs_path = file_path.resolve()

    # Build command with or without timeout wrapper
    if timeout_cmd:
        cmd = [
            timeout_cmd,
            str(timeout),
            soffice,
            "--headless",
            "--invisible",
            f"macro:///{LIBRARY_NAME}.RecalcMacro.RecalcAndSave",
            str(abs_path),
        ]
    else:
        cmd = [
            soffice,
            "--headless",
            "--invisible",
            f"macro:///{LIBRARY_NAME}.RecalcMacro.RecalcAndSave",
            str(abs_path),
        ]

    try:
        subprocess.run(cmd, check=True, capture_output=True, timeout=timeout + 10)
    except subprocess.TimeoutExpired:
        return {"status": "error", "message": "LibreOffice timed out"}
    except subprocess.CalledProcessError as e:
        return {"status": "error", "message": f"LibreOffice error: {e.stderr.decode()}"}

    return {"status": "success"}


def check_for_errors(file_path: Path):
    """Check Excel file for formula errors using openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed"}

    wb = load_workbook(file_path, data_only=True)
    errors = []
    formula_count = 0
    error_counts = {err: 0 for err in EXCEL_ERRORS}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    value_str = str(cell.value)
                    for error_type in EXCEL_ERRORS:
                        if error_type in value_str:
                            errors.append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "error": error_type,
                            })
                            error_counts[error_type] += 1

    # Also count formulas from non-data-only load
    wb_formulas = load_workbook(file_path, data_only=False)
    for sheet_name in wb_formulas.sheetnames:
        sheet = wb_formulas[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith("="):
                    formula_count += 1

    wb.close()
    wb_formulas.close()

    return {
        "status": "errors_found" if errors else "success",
        "total_errors": len(errors),
        "formula_count": formula_count,
        "error_counts": {k: v for k, v in error_counts.items() if v > 0},
        "errors": errors[:50],  # Limit to first 50 errors
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    file_path = Path(sys.argv[1])
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60

    if not file_path.exists():
        print(json.dumps({"status": "error", "message": f"File not found: {file_path}"}))
        sys.exit(1)

    # Recalculate formulas
    recalc_result = run_libreoffice_recalc(file_path, timeout)
    if recalc_result["status"] == "error":
        print(json.dumps(recalc_result))
        sys.exit(1)

    # Check for errors
    result = check_for_errors(file_path)
    print(json.dumps(result, indent=2))

    if result["status"] == "errors_found":
        sys.exit(1)


if __name__ == "__main__":
    main()
